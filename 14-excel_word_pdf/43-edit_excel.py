#!/usr/bin/env python
# pip install openpyxl
import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb["Sheet"]
sheet["A1"].value = 42
sheet["A2"].value = "Hello"

sheet2 = wb.create_sheet("Sheet2")
sheet.title = "Sheet1"
print(wb.sheetnames)

sheet3 = wb.create_sheet(index=0, title="New Sheet")
print(wb.sheetnames)

wb.save("test.xlsx")
