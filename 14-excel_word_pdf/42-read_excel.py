#!/usr/bin/env python
# pip install openpyxl
import openpyxl

workbook = openpyxl.load_workbook("example.xlsx")
print(f"workbook type={type(workbook)}")
print(f"sheets={workbook.sheetnames}")


sheet = workbook["Sheet1"]
print(f"sheet Sheet1 type={type(sheet)}")

cell = sheet["A1"]
print(f"cell A1 type={type(cell.value)}")
print(f"cell A1 value={str(cell.value)}")

print(f"cell B1 value={sheet['B1'].value}")
print(f"cell C1 value={sheet['C1'].value}")

print(f"Cell row 1, columne 2 value={sheet.cell(row=1, column=2).value}")

print("==============================")

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
    